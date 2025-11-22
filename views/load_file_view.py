import pandas as pd
import customtkinter as ctk
from views.ctk_dialogs import showinfo, showwarning, showerror, askstring, askinteger, askyesno
from styles import TkinterDialogStyles, AppColors, AppFonts, ButtonStyles
from openpyxl import load_workbook
import xlwt
import os
import tempfile
import csv
import chardet
import io
from styles import TkinterDialogStyles, AppColors, AppFonts, ButtonStyles

# Helper to deduplicate header names (Name, Name (1), Name (2) ...)
def _dedupe_headers(raw_headers):
    deduped = []
    counts = {}
    for name in raw_headers:
        base = (name or '').strip() or 'Unnamed'
        counts[base] = counts.get(base, 0) + 1
        if counts[base] == 1:
            deduped.append(base)
        else:
            deduped.append(f"{base} ({counts[base]-1})")
    return deduped

def truncate_text(value, max_length=20):
    """Truncate text if it exceeds the max length, adding '...' at the end."""
    if isinstance(value, str) and len(value) > max_length:
        return value[:max_length] + "..."
    return value

def format_dataframe(df, max_length=20):
    """Apply truncation to all text values in the DataFrame."""
    # DataFrame.applymap is deprecated in newer pandas — use DataFrame.map when available.
    try:
        # DataFrame.map applies elementwise mapping and avoids the applymap deprecation warning.
        return df.map(lambda x: truncate_text(str(x), max_length))
    except Exception:
        # Fallback for older pandas versions where map might not behave as expected.
        return df.applymap(lambda x: truncate_text(str(x), max_length))

def find_header_row(df, min_nonempty=None, pct_string_threshold=0.5):
    """Find a reasonable header row in a DataFrame.

    Heuristics used (in order):
    - Row has at least `min_nonempty` non-empty cells (default: max(3, 50% of columns)).
    - A majority of the non-empty cells in the row look like strings (not parseable as numbers).

    Returns the row index (integer) or None if not found.
    """
    if df is None or df.shape[0] == 0:
        return None

    ncols = df.shape[1]
    if min_nonempty is None:
        min_nonempty = max(3, int(0.5 * ncols))

    for i, row in df.iterrows():
        # Count non-empty values (not NaN and not empty string)
        non_empty_mask = ~row.isna() & (row.astype(str).str.strip() != "")
        non_empty_count = int(non_empty_mask.sum())
        if non_empty_count < min_nonempty:
            continue

        # Of the non-empty cells, how many are non-numeric when coerced?
        non_empty_values = row[non_empty_mask].astype(str).str.strip()
        coerced = pd.to_numeric(non_empty_values, errors='coerce')
        non_numeric_count = int(coerced.isna().sum())

        # If a sufficient fraction of the non-empty cells look like strings, treat as header
        if non_empty_count > 0 and (non_numeric_count / non_empty_count) >= pct_string_threshold:
            return i

    return None

def convert_xlsx_to_xls(xlsx_file, xls_file):
    wb_xlsx = load_workbook(xlsx_file, data_only=True)
    wb_xls = xlwt.Workbook()

    for sheet_name in wb_xlsx.sheetnames:
        sheet_xlsx = wb_xlsx[sheet_name]
        sheet_xls = wb_xls.add_sheet(sheet_name[:31])  # max 31 chars

        for row_idx, row in enumerate(sheet_xlsx.iter_rows()):
            for col_idx, cell in enumerate(row):
                if cell.value is not None:
                    sheet_xls.write(row_idx, col_idx, cell.value)

    wb_xls.save(xls_file)
    print(f"Converted: {xlsx_file} -> {xls_file}")


def convert_csv_to_xls(csv_file, xls_file, encoding='utf-8', delimiter=None):
    # Try multiple encodings with simple fallback, optionally sniff using chardet
    tried = []
    encodings_to_try = [encoding, 'utf-8-sig', 'cp1252', 'latin-1']
    raw_data = None
    for enc in encodings_to_try:
        if enc in tried:
            continue
        try:
            with open(csv_file, 'r', encoding=enc, newline='') as f:
                if raw_data is None:
                    raw_data = f.read(4096)
                    f.seek(0)
                # Auto-detect delimiter if not provided
                _sample = f.read(1024)
                f.seek(0)
                if delimiter is None:
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(_sample)
                        _delim = dialect.delimiter
                    except csv.Error:
                        _delim = ','
                else:
                    _delim = delimiter
                reader = csv.reader(f, delimiter=_delim)
                wb = xlwt.Workbook()
                ws = wb.add_sheet('Sheet1')
                for row_idx, row in enumerate(reader):
                    for col_idx, value in enumerate(row):
                        ws.write(row_idx, col_idx, value)
                wb.save(xls_file)
                print(f"Converted: {csv_file} -> {xls_file} (encoding={enc}, delimiter='{_delim}')")
                return
        except UnicodeDecodeError as ue:
            tried.append(enc)
            continue
        except Exception as e:
            print(f"CSV convert unexpected error with encoding {enc}: {e}")
            tried.append(enc)
            continue
    # If all failed, last resort: detect with chardet
    if raw_data is not None:
        detect = chardet.detect(raw_data.encode('latin-1', errors='ignore')) if isinstance(raw_data, str) else chardet.detect(raw_data)
        guessed = detect.get('encoding')
        if guessed and guessed not in tried:
            try:
                with open(csv_file, 'r', encoding=guessed, newline='') as f:
                    reader = csv.reader(f)
                    wb = xlwt.Workbook()
                    ws = wb.add_sheet('Sheet1')
                    for row_idx, row in enumerate(reader):
                        for col_idx, value in enumerate(row):
                            ws.write(row_idx, col_idx, value)
                    wb.save(xls_file)
                    print(f"Converted: {csv_file} -> {xls_file} (encoding={guessed} via chardet)")
                    return
            except Exception as e:
                print(f"Chardet attempt failed: {e}")
    raise UnicodeDecodeError('csv-decoder', b'', 0, 1, 'Failed to decode file with tried encodings')


def detect_csv_encoding_and_delimiter(csv_file, encodings=None, sample_size=8192):
    """Try to detect a sensible encoding and delimiter for a CSV file.

    Returns (encoding, delimiter). May return (None, None) if detection fails.
    """
    if encodings is None:
        encodings = ['utf-8', 'utf-8-sig', 'cp1252', 'latin-1']

    # Read a sample from the file as bytes for chardet
    with open(csv_file, 'rb') as f:
        raw = f.read(sample_size)

    detect = chardet.detect(raw)
    guessed = detect.get('encoding')

    # Try guessed encoding first, then fallbacks
    try_encodings = []
    if guessed:
        try_encodings.append(guessed)
    for e in encodings:
        if e not in try_encodings:
            try_encodings.append(e)

    sample_text = None
    for enc in try_encodings:
        try:
            sample_text = raw.decode(enc, errors='replace')
            break
        except Exception:
            continue

    if sample_text is None:
        return None, None

    # Try to sniff delimiter using csv.Sniffer
    sniffer = csv.Sniffer()
    delimiter = None
    try:
        dialect = sniffer.sniff(sample_text)
        delimiter = dialect.delimiter
    except csv.Error:
        # common fallback choices
        for d in [',', '\t', ';', '|']:
            if d in sample_text:
                delimiter = d
                break

    return try_encodings[0], delimiter

def load_file_view(file_path):
    """Load Excel file (.xls or .xlsx), convert if needed, and allow user to confirm header row."""
    # Ensure cache folder exists
    cache_dir = os.path.join(os.path.dirname(__file__), "file_cache")
    os.makedirs(cache_dir, exist_ok=True)

    # Clean up old cache files
    for f in os.listdir(cache_dir):
        fp = os.path.join(cache_dir, f)
        try:
            if os.path.isfile(fp):
                os.unlink(fp)
        except Exception as e:
            print(f"Failed to delete cache file {fp}: {e}")

    # Handle CSV directly where possible (avoids conversion to XLS which can change empty cells/encodings)
    converted_path = None
    if file_path.lower().endswith(".csv"):
        # Try detection and read with pandas.read_csv
        enc, delim = detect_csv_encoding_and_delimiter(file_path)
        try:
            if delim:
                df = pd.read_csv(file_path, header=None, encoding=enc or 'utf-8', sep=delim, engine='python')
            else:
                df = pd.read_csv(file_path, header=None, encoding=enc or 'utf-8')
        except Exception:
            # Fallback: convert to XLS and read via Excel reader if pandas CSV read fails
            converted_path = os.path.join(cache_dir, os.path.basename(file_path).replace(".csv", "_converted.xls"))
            convert_csv_to_xls(file_path, converted_path)
            file_path = converted_path
            df = pd.read_excel(file_path, engine="xlrd", header=None)
    elif file_path.lower().endswith(".xlsx"):
        converted_path = os.path.join(cache_dir, os.path.basename(file_path).replace(".xlsx", "_converted.xls"))
        convert_xlsx_to_xls(file_path, converted_path)
        file_path = converted_path
        df = pd.read_excel(file_path, engine="xlrd", header=None)
    else:
        # Assume Excel (.xls) or similar
        df = pd.read_excel(file_path, engine="xlrd", header=None)
    header_row = find_header_row(df)
    if header_row is None:
        showerror("Header Detection Failed", "Could not auto-detect a header row. Please verify the file format.")
        return None, None, None

    root = ctk.CTkToplevel()
    root.title("Header Row Preview")
    root.geometry("1200x700")
    root.configure(fg_color=TkinterDialogStyles.DIALOG_BG)

    # Ensure window is on top and modal
    root.lift()
    root.focus_force()
    root.grab_set()
    try:
        if hasattr(ctk, "_get_ancestor_window"):
            parent = ctk._get_ancestor_window()
            if parent:
                root.transient(parent)
    except Exception:
        pass

    df_truncated = format_dataframe(df, max_length=20)

    # Create frame with scrollbar
    frame = ctk.CTkFrame(root, fg_color=TkinterDialogStyles.FRAME_BG)
    frame.pack(pady=5, padx=5, fill="both", expand=True)

    # Use CTkTextbox with monospaced font for better alignment; colors from styles
    text_widget = ctk.CTkTextbox(frame, wrap="none", height=400, width=1100,
                      fg_color=TkinterDialogStyles.CANVAS_BG,
                      text_color=AppColors.BLACK,
                      font=("Courier New", 12))
    # Format with better alignment - show first 50 rows
    df_display = df_truncated.head(50)
    text_widget.insert("1.0", df_display.to_string(index=True))
    text_widget.configure(state="disabled")
    text_widget.pack(side="left", fill="both", expand=True)

    # Header selection
    header_frame = ctk.CTkFrame(root, fg_color=TkinterDialogStyles.FRAME_BG)
    header_frame.pack(pady=5)

    ctk.CTkLabel(header_frame, text="Header Row:",
                 text_color=AppColors.BLACK, font=AppFonts.BODY).grid(row=0, column=0, padx=5, sticky="w")
    header_input = ctk.CTkEntry(header_frame, width=TkinterDialogStyles.INPUT_WIDTH)
    header_input.insert(0, str(header_row))
    header_input.grid(row=0, column=1, padx=5, sticky="w")

    result = {"header_row": None, "keep_input": []}

    def on_confirm():
        try:
            user_header_row = int(header_input.get())
            selected_headers = _dedupe_headers(df.iloc[user_header_row].astype(str).tolist())  # Automatically select all columns

            result["header_row"] = user_header_row
            result["keep_input"] = selected_headers
            root.destroy()
        except ValueError:
            showerror("Invalid Input", "Please select valid header.")

    def on_cancel():
        root.destroy()

    # Buttons
    button_frame = ctk.CTkFrame(root, fg_color=TkinterDialogStyles.FRAME_BG)
    button_frame.pack(pady=10)

    ctk.CTkButton(button_frame, text="Confirm", command=on_confirm, **ButtonStyles.DEFAULT).grid(row=0, column=0, padx=10)
    ctk.CTkButton(button_frame, text="Cancel", command=on_cancel, **ButtonStyles.DEFAULT).grid(row=0, column=1, padx=10)

    root.wait_window()

    return df, result["header_row"], result["keep_input"]

